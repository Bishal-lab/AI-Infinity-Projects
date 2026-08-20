"""Configuration validation and the Excel management pack."""

from __future__ import annotations

import io

import openpyxl
import pytest
import yaml

from comms_dashboard.config import (
    load_all_mappings,
    load_settings,
    load_source_mapping,
    load_stage_ladder,
)
from comms_dashboard.export.excel import build_workbook
from comms_dashboard.metrics import FilterState
from comms_dashboard.models import STAGE_ORDER, ConfigError, SupportStatus


class TestStageLadder:
    def test_every_channel_declares_every_stage(self, ladder):
        """An omitted cell is ambiguous between 'n/a' and 'we forgot'."""
        for channel in ladder.channel_names:
            for stage in STAGE_ORDER:
                assert ladder.get(channel, stage) is not None

    def test_a_missing_stage_entry_is_rejected_with_a_useful_message(self, tmp_path):
        raw = yaml.safe_load(
            (load_stage_ladder.__globals__["CONFIG_DIR"] / "stage_ladder.yaml")
            .read_text(encoding="utf-8")
        )
        del raw["channels"]["email"]["stages"]["COMPLETED"]
        path = tmp_path / "ladder.yaml"
        path.write_text(yaml.safe_dump(raw), encoding="utf-8")

        with pytest.raises(ConfigError) as exc:
            load_stage_ladder(path)
        assert "COMPLETED" in str(exc.value)
        assert "email" in str(exc.value)

    def test_only_person_counting_stages_can_enter_funnel_maths(self, ladder):
        supporting = ladder.channels_supporting("ENGAGED")
        # Viva counts interactions, not people, so it must be excluded.
        assert "viva_engage" not in supporting
        assert "email" in supporting

    def test_completed_is_genuinely_unavailable_on_three_channels(self, ladder):
        na = [
            c for c in ladder.channel_names
            if ladder.get(c, "COMPLETED").status is SupportStatus.NOT_APPLICABLE
        ]
        assert set(na) == {"email", "whatsapp", "viva_engage"}

    def test_every_not_applicable_cell_explains_itself(self, ladder):
        """The UI shows these to management; "n/a" alone invites a bad guess."""
        for support in ladder.support.values():
            if support.status is SupportStatus.NOT_APPLICABLE:
                assert support.caveat, (support.channel, support.stage)


class TestSettings:
    def test_a_bad_date_order_is_rejected(self, tmp_path, settings):
        raw = dict(settings.raw)
        raw["parsing"] = {**raw["parsing"], "date_order": "YMD"}
        path = tmp_path / "settings.yaml"
        path.write_text(yaml.safe_dump(raw), encoding="utf-8")
        with pytest.raises(ConfigError, match="date_order"):
            load_settings(path)

    def test_a_bad_open_rate_basis_is_rejected(self, tmp_path, settings):
        raw = dict(settings.raw)
        raw["rules"] = {**raw.get("rules", {}), "open_rate_basis": "sent"}
        path = tmp_path / "settings.yaml"
        path.write_text(yaml.safe_dump(raw), encoding="utf-8")
        with pytest.raises(ConfigError, match="open_rate_basis"):
            load_settings(path)

    def test_a_missing_section_names_itself(self, tmp_path):
        path = tmp_path / "settings.yaml"
        path.write_text(yaml.safe_dump({"app": {}}), encoding="utf-8")
        with pytest.raises(ConfigError, match="paths"):
            load_settings(path)

    def test_thresholds_band_correctly(self, settings):
        threshold = settings.thresholds["engagement_rate"]
        assert threshold.rag(threshold.target + 0.01) == "green"
        assert threshold.rag(threshold.target * threshold.amber_factor) == "amber"
        assert threshold.rag(0.0) == "red"
        assert threshold.rag(None) == "grey"


class TestMappings:
    def test_every_mapping_has_a_channel_or_is_a_dimension(self, mappings, ladder):
        for source, mapping in mappings.items():
            if mapping.grain == "dimension":
                continue
            assert source in ladder.channel_names, source

    def test_every_mapping_declares_detection_rules(self, mappings):
        for source, mapping in mappings.items():
            assert mapping.detect.strong or mapping.detect.filename_glob, source
            assert mapping.detect.min_score > 0, source

    def test_a_duplicate_source_across_files_is_rejected(self, tmp_path, mappings):
        directory = tmp_path / "mappings"
        directory.mkdir()
        for name in ("a.yaml", "b.yaml"):
            (directory / name).write_text(
                yaml.safe_dump({"source": "email", "fields": {}}), encoding="utf-8"
            )
        with pytest.raises(ConfigError, match="duplicate source"):
            load_all_mappings(directory)

    def test_a_missing_source_key_names_the_file(self, tmp_path):
        path = tmp_path / "broken.yaml"
        path.write_text(yaml.safe_dump({"fields": {}}), encoding="utf-8")
        with pytest.raises(ConfigError, match="broken.yaml.*source"):
            load_source_mapping(path)


class TestExcelPack:
    @pytest.fixture
    def workbook(self, loaded, settings):
        payload, filename = build_workbook(loaded, FilterState(), 1, settings)
        return openpyxl.load_workbook(io.BytesIO(payload)), filename

    def test_it_contains_the_expected_sheets(self, workbook):
        book, _ = workbook
        assert set(book.sheetnames) >= {
            "Cover", "Summary KPIs", "Funnel by channel", "Campaign detail",
            "Coverage", "Data quality", "Definitions",
        }

    def test_unmeasurable_stages_are_words_not_zeroes(self, workbook):
        book, _ = workbook
        sheet = book["Funnel by channel"]
        values = [
            row[3] for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True)
            if row[0]
        ]
        assert "n/a" in values
        # If any of them had become 0 the workbook would be quietly wrong.
        assert not any(v == 0 for v in values)

    def test_it_carries_no_person_level_rows(self, workbook):
        """Forwarding this to a leadership team must not leak an individual."""
        book, _ = workbook
        for name in book.sheetnames:
            for row in book[name].iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str):
                        assert "@" not in cell or "example.com" not in cell
                        assert not cell.startswith("+44")

    def test_the_cover_states_the_view_it_was_built_from(self, workbook):
        book, _ = workbook
        text = " ".join(
            str(c) for row in book["Cover"].iter_rows(values_only=True)
            for c in row if c
        )
        assert "all campaigns" in text
        assert "n/a" in text  # the how-to-read-blanks note
